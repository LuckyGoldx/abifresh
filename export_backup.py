#!/usr/bin/env python3
"""
Export all Supabase tables via backend API
"""
import requests
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import json

# Backupfile should be created with these tables
TABLES = [
    'users',
    'items',
    'inventory_main_store',
    'inventory_active_store',
    'sales',
    'receipts',
    'staff_payments',
    'staff_expenses',
    'system_settings',
]

print(f"🔄 Fetching {len(TABLES)} tables via API...")

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

exported_count = 0
base_url = "http://localhost:5000/api/backup/table"

for table in TABLES:
    try:
        # Fetch all rows from API
        url = f"{base_url}/{table}/all"
        response = requests.get(url, timeout=10)
        
        if response.status_code != 200:
            print(f"  ⚠️  {table}: API error {response.status_code}")
            continue
        
        data = response.json()
        rows = data.get('rows', [])
        
        if rows:
            df = pd.DataFrame(rows)
            ws = wb.create_sheet(title=table)
            
            # Write headers
            for col_num, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=col_name)
            
            # Write data
            for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            print(f"  ✅ {table}: {len(rows)} rows")
            exported_count += 1
        else:
            print(f"  ⏭️  {table}: 0 rows (skipped)")
    except Exception as e:
        print(f"  ❌ {table}: {str(e)}")

# Save
filename = f"backup_current_state_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)
print(f"\n✅ Exported {exported_count} tables to: {filename}")

