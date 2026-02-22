"""
Generate comprehensive Hospital EMR Pitch Document in Excel format.
Multi-sheet workbook with formatted tables, diagrams, and charts.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "Hospital_EMR_Pitch")
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ─── Color Palette ────────────────────────────────────────────────────
DARK_BLUE = "003366"
MED_BLUE = "006699"
LIGHT_BLUE = "E6F0FA"
WHITE = "FFFFFF"
DARK_GREEN = "006633"
LIGHT_GREEN = "E6F5E6"
DARK_RED = "990000"
LIGHT_RED = "FFE6E6"
LIGHT_YELLOW = "FFFDE6"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
DARK_GRAY = "333333"
ACCENT_ORANGE = "E67300"
ACCENT_PURPLE = "660066"

# ─── Reusable Styles ─────────────────────────────────────────────────
title_font = Font(name="Calibri", size=20, bold=True, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", size=14, bold=True, color=MED_BLUE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
subheader_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
alt_row_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_title(ws, row, col, text, merge_end_col=None):
    """Write a title cell with formatting."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if merge_end_col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=merge_end_col
        )
    return row + 1


def write_subtitle(ws, row, col, text, merge_end_col=None):
    """Write a subtitle cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if merge_end_col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=merge_end_col
        )
    return row + 1


def write_table(ws, start_row, headers, rows, col_start=1, col_widths=None):
    """Write a formatted table with headers and alternating row colors."""
    # Headers
    for c_idx, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=col_start + c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for r_idx, row_data in enumerate(rows):
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(
                row=start_row + 1 + r_idx,
                column=col_start + c_idx,
                value=val,
            )
            cell.font = normal_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = alt_row_fill

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(col_start + i)].width = w

    return start_row + 1 + len(rows) + 1  # next available row


def write_paragraph(ws, row, col, text, merge_end_col=None):
    """Write a paragraph of text."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = normal_font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if merge_end_col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=merge_end_col
        )
    return row + 1


# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: COVER PAGE
# ═══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Cover Page"
ws.sheet_properties.tabColor = DARK_BLUE

# Set column widths
for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 18

ws.row_dimensions[8].height = 50
ws.row_dimensions[10].height = 30
ws.row_dimensions[12].height = 25

ws.merge_cells("B8:F8")
cell = ws["B8"]
cell.value = "🏥 HOSPITAL EMR SYSTEM"
cell.font = Font(name="Calibri", size=28, bold=True, color=DARK_BLUE)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B10:F10")
cell = ws["B10"]
cell.value = "Electronic Medical Records & Hospital Management Platform"
cell.font = Font(name="Calibri", size=16, color=MED_BLUE)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B12:F12")
cell = ws["B12"]
cell.value = "Comprehensive Solution for Private Hospital Operations"
cell.font = Font(name="Calibri", size=13, italic=True, color=DARK_GRAY)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B16:F16")
cell = ws["B16"]
cell.value = "CONFIDENTIAL PITCH DOCUMENT"
cell.font = Font(name="Calibri", size=12, bold=True, color=DARK_RED)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B18:F18")
cell = ws["B18"]
cell.value = "Prepared: February 2026  |  Version 1.0"
cell.font = Font(name="Calibri", size=11, color=DARK_GRAY)
cell.alignment = Alignment(horizontal="center", vertical="center")

# Add a blue banner
for col in range(2, 7):
    ws.cell(row=6, column=col).fill = PatternFill(
        start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid"
    )
    ws.cell(row=20, column=col).fill = PatternFill(
        start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid"
    )

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Executive Summary")
ws2.sheet_properties.tabColor = MED_BLUE

for col in range(1, 5):
    ws2.column_dimensions[get_column_letter(col)].width = [5, 30, 50, 5][col - 1]

r = 2
r = write_title(ws2, r, 2, "EXECUTIVE SUMMARY", merge_end_col=3)
r += 1

r = write_paragraph(
    ws2, r, 2,
    "Our Hospital EMR system is a comprehensive, integrated healthcare management platform "
    "designed specifically for private hospitals. It digitizes every aspect of hospital operations — "
    "from patient registration to discharge, from pharmacy dispensing to laboratory results, "
    "from staff scheduling to financial reporting.",
    merge_end_col=3,
)
ws2.row_dimensions[r - 1].height = 60
r += 1

r = write_subtitle(ws2, r, 2, "Key Value Propositions", merge_end_col=3)
r += 1

benefits = [
    ("✅ Paperless Operations", "Eliminate paper records, forms, and manual tracking entirely"),
    ("✅ Improved Patient Safety", "Drug interaction alerts, allergy warnings, clinical decision support"),
    ("✅ Revenue Optimization", "Automated billing, insurance claims, and financial analytics"),
    ("✅ Operational Efficiency", "Streamlined workflows reduce wait times by up to 60%"),
    ("✅ Regulatory Compliance", "Full audit trails, data encryption, role-based access control"),
    ("✅ Data-Driven Decisions", "Real-time dashboards, KPIs, and business intelligence reports"),
    ("✅ Staff Productivity", "Automated scheduling, task assignment, and workload balancing"),
    ("✅ Patient Satisfaction", "Online booking, patient portal, SMS/email notifications"),
]
r = write_table(ws2, r, ["Benefit", "Description"], benefits, col_start=2, col_widths=[30, 50])
r += 1

r = write_subtitle(ws2, r, 2, "ROI Highlights", merge_end_col=3)
r += 1

roi = [
    ("Wait Time Reduction", "Up to 60% decrease in patient wait times"),
    ("Revenue Increase", "15-25% improvement in revenue capture"),
    ("Error Reduction", "90% fewer medication and transcription errors"),
    ("Staff Productivity", "30% increase in staff efficiency"),
    ("Paper Cost Savings", "Eliminate $50K-$200K annual paper & storage costs"),
    ("Insurance Claims", "40% faster claim processing and reimbursement"),
]
r = write_table(ws2, r, ["Metric", "Impact"], roi, col_start=2, col_widths=[30, 50])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: SYSTEM ARCHITECTURE
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Architecture")
ws3.sheet_properties.tabColor = DARK_GREEN

for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = [5, 25, 25, 35, 5][col - 1]

r = 2
r = write_title(ws3, r, 2, "SYSTEM ARCHITECTURE", merge_end_col=4)
r += 1

# Architecture layers
r = write_subtitle(ws3, r, 2, "Three-Tier Architecture", merge_end_col=4)
r += 1

arch_layers = [
    ("Presentation Layer", "React.js / Next.js / React Native", "Web dashboard, mobile apps, patient portal, self-service kiosks"),
    ("Application Layer", "Node.js (NestJS) / Python (FastAPI)", "REST APIs, GraphQL, WebSocket, business logic, RBAC"),
    ("Data Layer", "PostgreSQL + Redis + Elasticsearch", "Relational data, caching, full-text search, file storage"),
]
r = write_table(ws3, r, ["Layer", "Technology", "Purpose"], arch_layers, col_start=2, col_widths=[25, 25, 35])
r += 1

# Component details
r = write_subtitle(ws3, r, 2, "System Components", merge_end_col=4)
r += 1

components = [
    ("Frontend (Web)", "React.js / Next.js 14", "Server-side rendering, responsive UI, PWA"),
    ("Mobile Application", "React Native / Flutter", "Cross-platform iOS & Android, offline mode"),
    ("API Gateway", "NGINX / Traefik", "Load balancing, rate limiting, SSL termination"),
    ("Auth Service", "JWT + OAuth 2.0", "SSO, MFA, session management"),
    ("Core EMR API", "Node.js / Python", "Patient, clinical, orders, prescriptions"),
    ("Analytics Service", "Python + Pandas", "Reports, dashboards, BI"),
    ("Database", "PostgreSQL 16+", "ACID compliance, JSON support, partitioning"),
    ("Cache", "Redis", "Session store, real-time data, pub/sub"),
    ("Search", "Elasticsearch", "Patient search, clinical search"),
    ("File Storage", "MinIO / S3", "Medical images, documents, DICOM"),
    ("Message Queue", "RabbitMQ", "Async processing, notifications"),
    ("PACS Server", "Orthanc", "DICOM image storage and viewing"),
    ("Monitoring", "Prometheus + Grafana", "System health, performance metrics"),
    ("Logging", "ELK Stack", "Centralized log management"),
    ("CI/CD", "GitHub Actions", "Automated testing and deployment"),
]
r = write_table(ws3, r, ["Component", "Technology", "Purpose"], components, col_start=2, col_widths=[25, 25, 35])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: CORE MODULES
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Core Modules")
ws4.sheet_properties.tabColor = "006699"

for col in range(1, 6):
    ws4.column_dimensions[get_column_letter(col)].width = [5, 22, 28, 40, 5][col - 1]

r = 2
r = write_title(ws4, r, 2, "CORE MODULES & FEATURES", merge_end_col=4)
r += 1

# Module overview
modules_overview = [
    ("Patient Management", "Clinical", "Patient registration, profiles, UHID, search, portal, communications"),
    ("Electronic Medical Records", "Clinical", "SOAP notes, ICD coding, e-prescribing, vitals, orders, clinical decision support"),
    ("Appointment & Scheduling", "Clinical", "Online booking, walk-in queue, doctor schedules, telemedicine, reminders"),
    ("OPD Management", "Clinical", "Check-in, triage, doctor queue, consultation, referrals, billing"),
    ("In-Patient (Ward) Mgmt", "Clinical", "Admission, bed management, transfers, diet, discharge planning"),
    ("Emergency Department", "Clinical", "Rapid registration, ESI triage, trauma, resuscitation, disposition"),
    ("Operating Theatre", "Clinical", "Surgery scheduling, WHO checklist, OT docs, implant tracking"),
    ("Nursing Station", "Clinical", "Worklist, vitals charting, eMAR, care plans, shift handover"),
    ("Laboratory (LIS)", "Support", "Order management, sample tracking, results, QC, instrument interface"),
    ("Radiology (RIS/PACS)", "Support", "Imaging orders, DICOM storage, web viewer, structured reports"),
    ("Pharmacy", "Support", "E-prescriptions, dispensing, drug interactions, inventory, POS"),
    ("Blood Bank", "Support", "Donors, collection, cross-match, transfusions, testing, compliance"),
    ("Inventory & Supply", "Support", "Multi-store, purchase orders, GRN, expiry, ABC/VED analysis"),
    ("Billing & Finance", "Financial", "Patient billing, insurance claims, payments, deposits, accounting"),
    ("Staff & HR", "Administrative", "Employee profiles, attendance, shifts, payroll, performance, leave"),
    ("Admin Dashboard", "Administrative", "KPIs, financial overview, occupancy metrics, staff overview"),
    ("Reports & BI", "Administrative", "Clinical, financial, operational, custom report builder"),
    ("User & Role Mgmt", "Administrative", "RBAC, roles, permissions, MFA, SSO, audit trail"),
]
r = write_table(
    ws4, r,
    ["Module", "Category", "Key Features"],
    modules_overview,
    col_start=2,
    col_widths=[22, 28, 40],
)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5: PATIENT MANAGEMENT DETAILS
# ═══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Patient Management")
ws5.sheet_properties.tabColor = "006633"

for col in range(1, 5):
    ws5.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws5, r, 2, "PATIENT MANAGEMENT MODULE", merge_end_col=3)
r += 1

r = write_paragraph(
    ws5, r, 2,
    "Complete patient lifecycle management from first contact through discharge and follow-up.",
    merge_end_col=3,
)
r += 1

patient_features = [
    ("Patient Registration", "Demographics, contact info, emergency contacts, photo capture, biometric ID"),
    ("Unique Patient ID (UHID)", "Auto-generated unique hospital ID with barcode/QR code"),
    ("Patient Search", "Quick search by name, ID, phone, national ID — instant results"),
    ("Patient Categories", "VIP, regular, insurance, corporate, staff, pediatric, geriatric"),
    ("Medical History", "Past illnesses, surgeries, family history, social history, allergies"),
    ("Insurance Management", "Multiple insurance policies, pre-authorization, coverage limits"),
    ("Document Upload", "ID cards, insurance cards, referral letters, consent forms"),
    ("Patient Portal", "Self-service: view records, book appointments, download reports"),
    ("Patient Merge/Link", "Merge duplicate records, link family members"),
    ("Communication", "SMS, email, WhatsApp notifications for appointments & results"),
]
r = write_table(ws5, r, ["Feature", "Description"], patient_features, col_start=2, col_widths=[25, 55])
r += 1

# Registration workflow
r = write_subtitle(ws5, r, 2, "Patient Registration Workflow", merge_end_col=3)
r += 1

workflow_steps = [
    ("Step 1", "Patient Arrives at Hospital (walk-in, appointment, or emergency)"),
    ("Step 2", "Registration Desk / Self-Service Kiosk check-in"),
    ("Step 3", "Search existing records by name, phone, or national ID"),
    ("Step 4", "New patient? Create profile and generate UHID"),
    ("Step 5", "Capture demographics, insurance details, emergency contacts"),
    ("Step 6", "Issue patient card with QR/barcode"),
    ("Step 7", "Direct patient to appropriate department (OPD, ER, Lab)"),
]
r = write_table(ws5, r, ["Step", "Action"], workflow_steps, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6: EMR DETAILS
# ═══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("EMR Details")
ws6.sheet_properties.tabColor = "003399"

for col in range(1, 5):
    ws6.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws6, r, 2, "ELECTRONIC MEDICAL RECORDS", merge_end_col=3)
r += 1

emr_features = [
    ("Clinical Notes", "SOAP notes, progress notes, consultation notes with specialty templates"),
    ("Diagnosis (ICD-10/11)", "Searchable ICD coding, multiple diagnoses per encounter"),
    ("E-Prescribing", "Drug interaction checks, dosage calculators, generic substitution"),
    ("Vitals Recording", "BP, temperature, pulse, SpO2, weight, height, BMI auto-calculation"),
    ("Order Management", "Lab orders, radiology orders, procedure orders — all digital"),
    ("Clinical Templates", "Specialty-specific templates (cardiology, orthopedics, OB/GYN, etc.)"),
    ("Allergy & Alert System", "Drug allergies, food allergies, condition alerts — prominent display"),
    ("Document Generation", "Auto-generate discharge summaries, referral letters, certificates"),
    ("Medical Timeline", "Chronological view of all patient encounters, results, procedures"),
    ("Clinical Decision Support", "Evidence-based alerts, guideline reminders, risk scoring"),
    ("Voice-to-Text", "Dictation support for clinical notes with AI transcription"),
    ("E-Signatures", "Digital signatures for prescriptions, consent forms, documents"),
]
r = write_table(ws6, r, ["Feature", "Description"], emr_features, col_start=2, col_widths=[25, 55])
r += 1

# Clinical workflow
r = write_subtitle(ws6, r, 2, "Clinical Documentation Workflow", merge_end_col=3)
r += 1

clinical_workflow = [
    ("Step 1", "Patient encounter begins — doctor opens patient record via UHID"),
    ("Step 2", "Review medical history, allergies, and active alerts"),
    ("Step 3", "Record vital signs and chief complaint"),
    ("Step 4", "Clinical examination and SOAP note documentation"),
    ("Step 5", "Order labs, imaging, or procedures as needed"),
    ("Step 6", "Prescribe medications with drug interaction safety check"),
    ("Step 7", "Generate encounter summary and treatment plan"),
    ("Step 8", "Schedule follow-up appointment or initiate admission"),
]
r = write_table(ws6, r, ["Step", "Action"], clinical_workflow, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 7: SUPPORT MODULES (Lab, Pharmacy, Radiology)
# ═══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Lab-Pharmacy-Radiology")
ws7.sheet_properties.tabColor = ACCENT_PURPLE

for col in range(1, 5):
    ws7.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws7, r, 2, "SUPPORT MODULES", merge_end_col=3)
r += 1

# Laboratory
r = write_subtitle(ws7, r, 2, "🔬 Laboratory Information System (LIS)", merge_end_col=3)
r += 1

lab_features = [
    ("Order Management", "Receive lab orders digitally, priority flagging (STAT/routine)"),
    ("Sample Collection", "Barcode labels, phlebotomy scheduling, collection tracking"),
    ("Sample Tracking", "Track sample: collection → processing → reporting"),
    ("Result Entry", "Manual or instrument-interfaced entry with normal ranges"),
    ("Auto-Validation", "Rules-based auto-validation for results within normal ranges"),
    ("Critical Alerts", "Immediate notification for critical/panic values"),
    ("Report Generation", "Professional reports with reference ranges & interpretations"),
    ("Test Catalog", "Comprehensive test catalog, panels, profiles, pricing"),
    ("Quality Control", "QC tracking, Levey-Jennings charts, Westgard rules"),
    ("Instrument Interface", "Bi-directional connection with lab analyzers (ASTM/HL7)"),
    ("TAT Monitoring", "Turnaround time tracking and bottleneck identification"),
    ("External Referral", "Send samples to reference labs, track results"),
]
r = write_table(ws7, r, ["Feature", "Description"], lab_features, col_start=2, col_widths=[25, 55])
r += 1

# Pharmacy
r = write_subtitle(ws7, r, 2, "💊 Pharmacy Management", merge_end_col=3)
r += 1

pharmacy_features = [
    ("E-Prescription Receipt", "Receive prescriptions electronically from EMR"),
    ("Drug Dispensing", "Barcode-based dispensing, patient verification, label printing"),
    ("Drug Interaction Check", "Real-time alerts for drug-drug & drug-allergy interactions"),
    ("Inventory Management", "Stock levels, reorder points, expiry tracking, batch management"),
    ("Formulary Management", "Drug formulary with generic substitution suggestions"),
    ("Controlled Substances", "Special tracking for narcotics/controlled medications"),
    ("Purchase Orders", "Auto-generate POs when stock falls below reorder level"),
    ("Supplier Management", "Vendor database, price comparison, delivery tracking"),
    ("Point of Sale", "Walk-in customer sales, OTC medications, billing"),
    ("Reports", "Fast/slow movers, expiry reports, revenue by category"),
]
r = write_table(ws7, r, ["Feature", "Description"], pharmacy_features, col_start=2, col_widths=[25, 55])
r += 1

# Radiology
r = write_subtitle(ws7, r, 2, "📷 Radiology & Imaging (RIS/PACS)", merge_end_col=3)
r += 1

radiology_features = [
    ("Order Management", "Digital radiology orders with clinical indication"),
    ("Scheduling", "Imaging appointment scheduling, equipment slot management"),
    ("DICOM Integration", "Modality Worklist, image storage, PACS server"),
    ("Web-Based Viewer", "DICOM viewer with measurement and comparison tools"),
    ("Structured Reporting", "Reporting templates, voice dictation support"),
    ("Report Distribution", "Auto-send reports to EMR, doctor, patient portal"),
    ("Equipment Tracking", "Maintenance schedules, utilization reports"),
]
r = write_table(ws7, r, ["Feature", "Description"], radiology_features, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 8: BILLING & FINANCE
# ═══════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Billing & Finance")
ws8.sheet_properties.tabColor = DARK_GREEN

for col in range(1, 5):
    ws8.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws8, r, 2, "BILLING & FINANCIAL MANAGEMENT", merge_end_col=3)
r += 1

billing_features = [
    ("Patient Billing", "Auto-generate bills from orders, procedures, bed charges, consumables"),
    ("Insurance Claims", "Electronic claim submission, pre-authorization, adjudication tracking"),
    ("Multiple Payments", "Cash, card, mobile money, bank transfer, installment plans"),
    ("Price Lists / Tariffs", "Configurable price lists for services, by department or package"),
    ("Package Billing", "Health check packages, surgery packages, maternity packages"),
    ("Deposit Management", "Advance deposits, refunds, deposit adjustment tracking"),
    ("Credit Management", "Corporate accounts, credit limits, outstanding tracking"),
    ("Revenue Dashboard", "Daily/weekly/monthly revenue, department-wise collection"),
    ("Tax Management", "VAT/GST calculation, tax invoices, exemptions"),
    ("Refund Processing", "Structured refund approval workflow with audit trail"),
    ("Financial Reports", "P&L, balance sheet, aging reports, revenue forecasts"),
    ("Accounting Integration", "Export to QuickBooks, Xero, Tally, or custom ERP"),
]
r = write_table(ws8, r, ["Feature", "Description"], billing_features, col_start=2, col_widths=[25, 55])
r += 1

# Billing workflow
r = write_subtitle(ws8, r, 2, "Billing Workflow", merge_end_col=3)
r += 1

billing_workflow = [
    ("Step 1", "Services rendered — consultation, lab, procedure, pharmacy"),
    ("Step 2", "Charges automatically captured to patient bill"),
    ("Step 3", "Insurance coverage check and pre-authorization if needed"),
    ("Step 4", "Generate itemized invoice for patient"),
    ("Step 5", "Patient payment — cash, card, mobile money, or installment"),
    ("Step 6", "Insurance claim submitted electronically"),
    ("Step 7", "Payment reconciliation and posting"),
    ("Step 8", "Receipt generated, financial reports updated"),
]
r = write_table(ws8, r, ["Step", "Action"], billing_workflow, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 9: STAFF & HR
# ═══════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Staff & HR")
ws9.sheet_properties.tabColor = ACCENT_ORANGE

for col in range(1, 5):
    ws9.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws9, r, 2, "STAFF & HR MANAGEMENT", merge_end_col=3)
r += 1

hr_features = [
    ("Employee Profiles", "Complete profiles with credentials, certifications, specializations"),
    ("Recruitment Pipeline", "Job postings, applications, interview scheduling, offers"),
    ("Attendance & Time", "Clock-in/out, biometric integration, overtime tracking"),
    ("Shift Scheduling", "Auto-generate shift rosters, swap requests, on-call management"),
    ("Leave Management", "Apply, approve, track leave — annual, sick, maternity, emergency"),
    ("Payroll Integration", "Salary calculation, deductions, bonuses, pay slip generation"),
    ("Performance Reviews", "360° evaluations, KPIs, competency assessments"),
    ("Training & CME", "Track continuing medical education, certifications, expiry alerts"),
    ("Credentialing", "Verify and track medical licenses, board certifications"),
    ("Department Mgmt", "Organizational hierarchy, department heads, reporting lines"),
    ("Staff Communication", "Internal messaging, announcements, policy distribution"),
    ("Expense Management", "Travel claims, reimbursements, petty cash management"),
]
r = write_table(ws9, r, ["Feature", "Description"], hr_features, col_start=2, col_widths=[25, 55])
r += 1

# Roles
r = write_subtitle(ws9, r, 2, "System Roles & Access", merge_end_col=3)
r += 1

roles = [
    ("Administrator", "Full system access, user management, configuration"),
    ("Doctor", "Patient records, prescriptions, orders, clinical notes"),
    ("Nurse", "Vitals, medication administration, nursing notes, care plans"),
    ("Receptionist", "Patient registration, appointments, billing"),
    ("Pharmacist", "Drug dispensing, inventory, prescription verification"),
    ("Lab Technician", "Sample processing, result entry, quality control"),
    ("Radiologist", "Image review, report dictation, worklist management"),
    ("Accountant", "Billing, insurance claims, financial reports"),
    ("HR Manager", "Staff management, payroll, attendance, recruitment"),
    ("Custom Roles", "Create any role with granular permission configuration"),
]
r = write_table(ws9, r, ["Role", "Access Scope"], roles, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 10: WORKFLOWS
# ═══════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Workflows")
ws10.sheet_properties.tabColor = "990066"

for col in range(1, 5):
    ws10.column_dimensions[get_column_letter(col)].width = [5, 12, 65, 5][col - 1]

r = 2
r = write_title(ws10, r, 2, "KEY WORKFLOWS", merge_end_col=3)
r += 1

# Patient Journey
r = write_subtitle(ws10, r, 2, "1. Complete Patient Journey", merge_end_col=3)
r += 1

patient_journey = [
    ("1", "Patient arrives at hospital (walk-in, scheduled appointment, or emergency)"),
    ("2", "Registration: new patient creates profile; returning patient verified by UHID"),
    ("3", "Triage / queue assignment based on urgency and department"),
    ("4", "Consultation: doctor reviews history, examines patient, documents SOAP note"),
    ("5", "Orders placed: lab tests, imaging, procedures, medications"),
    ("6", "Lab/Radiology: samples collected, processed, results sent to EMR"),
    ("7", "Pharmacy: e-prescription dispensed with safety checks"),
    ("8", "Decision: discharge with follow-up OR admit to ward"),
    ("9", "If admitted: bed assigned, daily rounds, nursing care, monitoring"),
    ("10", "Discharge: summary generated, medications dispensed, follow-up scheduled"),
    ("11", "Billing: all charges consolidated, payment collected, insurance claim filed"),
    ("12", "Post-discharge: follow-up reminders, patient satisfaction survey"),
]
r = write_table(ws10, r, ["Step", "Action"], patient_journey, col_start=2, col_widths=[12, 65])
r += 1

# Emergency workflow
r = write_subtitle(ws10, r, 2, "2. Emergency Department Workflow", merge_end_col=3)
r += 1

er_flow = [
    ("1", "Patient arrives via ambulance or walk-in — rapid registration (minimal info)"),
    ("2", "Triage assessment: ESI Level 1 (Resuscitation) to Level 5 (Non-urgent)"),
    ("3", "Assign to ER bay/bed based on acuity level"),
    ("4", "STAT labs and imaging ordered with priority flags"),
    ("5", "Emergency treatment initiated — medications, procedures, monitoring"),
    ("6", "Disposition decision: admit to ward, discharge, transfer, or observe"),
    ("7", "Documentation completed, handover to receiving department"),
    ("8", "ER dashboard updated with real-time metrics"),
]
r = write_table(ws10, r, ["Step", "Action"], er_flow, col_start=2, col_widths=[12, 65])
r += 1

# Surgery workflow
r = write_subtitle(ws10, r, 2, "3. Operating Theatre Workflow", merge_end_col=3)
r += 1

ot_flow = [
    ("1", "Surgery scheduled — OT slot booked, team assigned, equipment reserved"),
    ("2", "Pre-operative assessment: anesthesia evaluation, lab results reviewed"),
    ("3", "Consent forms signed digitally by patient and surgeon"),
    ("4", "WHO Surgical Safety Checklist — Sign-In phase"),
    ("5", "WHO Checklist — Time-Out: verify patient, procedure, site"),
    ("6", "Surgery performed — intra-operative documentation"),
    ("7", "WHO Checklist — Sign-Out: instruments counted, specimen labeled"),
    ("8", "Post-operative handover to recovery/PACU"),
    ("9", "Surgical notes, implant registry, consumables documented"),
    ("10", "Patient transferred to ward for post-operative care"),
]
r = write_table(ws10, r, ["Step", "Action"], ot_flow, col_start=2, col_widths=[12, 65])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 11: SECURITY & COMPLIANCE
# ═══════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Security")
ws11.sheet_properties.tabColor = DARK_RED

for col in range(1, 5):
    ws11.column_dimensions[get_column_letter(col)].width = [5, 25, 55, 5][col - 1]

r = 2
r = write_title(ws11, r, 2, "SECURITY & COMPLIANCE", merge_end_col=3)
r += 1

security = [
    ("Data Encryption", "AES-256 at rest, TLS 1.3 in transit — all data encrypted"),
    ("Access Control", "Role-based (RBAC), attribute-based (ABAC), IP whitelisting"),
    ("Multi-Factor Auth", "SMS OTP, authenticator app, biometric as second factor"),
    ("Session Management", "Auto-timeout, concurrent session limits, device tracking"),
    ("Audit Logging", "Every action logged: who, what, when, where, before/after values"),
    ("Data Backup", "Hourly incremental, daily full, off-site replication"),
    ("Disaster Recovery", "RPO < 1 hour, RTO < 4 hours, automated failover"),
    ("Penetration Testing", "Regular third-party security assessments"),
    ("HIPAA Compliance", "Full compliance with health data protection requirements"),
    ("GDPR Support", "Right to access, right to erasure, data portability"),
    ("SOC 2 Type II", "Service organization security controls certification"),
    ("Data Residency", "Choose data storage location for regulatory compliance"),
]
r = write_table(ws11, r, ["Security Feature", "Details"], security, col_start=2, col_widths=[25, 55])
r += 1

# Integrations
r = write_subtitle(ws11, r, 2, "Integration Standards", merge_end_col=3)
r += 1

integrations = [
    ("HL7 / FHIR", "Healthcare data exchange standards for interoperability"),
    ("DICOM", "Medical imaging standard for radiology equipment"),
    ("ICD-10 / ICD-11", "International classification of diseases for coding"),
    ("SNOMED CT", "Standardized clinical terminology"),
    ("Lab Instruments", "Bi-directional interface (ASTM/HL7)"),
    ("Insurance Portals", "Electronic claims, pre-auth, eligibility"),
    ("Payment Gateways", "Stripe, PayStack, Flutterwave, bank integrations"),
    ("SMS/Email Gateways", "Twilio, SendGrid for communications"),
    ("Video Conferencing", "Zoom, Jitsi for telemedicine"),
    ("Accounting Software", "QuickBooks, Xero, Tally integration"),
]
r = write_table(ws11, r, ["Integration", "Purpose"], integrations, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 12: IMPLEMENTATION PLAN
# ═══════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("Implementation")
ws12.sheet_properties.tabColor = ACCENT_ORANGE

for col in range(1, 7):
    ws12.column_dimensions[get_column_letter(col)].width = [5, 20, 12, 12, 40, 5][col - 1]

r = 2
r = write_title(ws12, r, 2, "IMPLEMENTATION PLAN", merge_end_col=5)
r += 1

r = write_paragraph(
    ws12, r, 2,
    "Phased implementation over 12 months ensures minimal disruption while delivering value incrementally.",
    merge_end_col=5,
)
r += 1

impl_phases = [
    ("Phase 1: Foundation", "Month 1-3", "3 months",
     "Project setup, infrastructure, patient registration, user management, basic EMR, appointment scheduling"),
    ("Phase 2: Clinical Core", "Month 4-6", "3 months",
     "Full EMR with templates, OPD workflow, laboratory (LIS), pharmacy management, billing, nursing station"),
    ("Phase 3: Advanced Clinical", "Month 7-9", "3 months",
     "In-patient management, emergency dept, operating theatre, radiology/PACS, blood bank, diet/kitchen"),
    ("Phase 4: Analytics", "Month 10-11", "2 months",
     "Admin dashboard & KPIs, business intelligence, report builder, insurance & payment integration"),
    ("Phase 5: Launch", "Month 12", "1 month",
     "Performance optimization, security hardening, user training, data migration, go-live support"),
]
r = write_table(
    ws12, r,
    ["Phase", "Timeline", "Duration", "Deliverables"],
    impl_phases,
    col_start=2,
    col_widths=[20, 12, 12, 40],
)
r += 1

# Gantt-style view
r = write_subtitle(ws12, r, 2, "Timeline Overview (Gantt)", merge_end_col=5)
r += 1

# Create simple gantt-style cells
months = ["", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"]
for c_idx, m in enumerate(months):
    cell = ws12.cell(row=r, column=2 + c_idx, value=m)
    cell.font = header_font if c_idx > 0 else bold_font
    cell.fill = header_fill if c_idx > 0 else PatternFill()
    cell.alignment = center_align
    cell.border = thin_border
    if c_idx > 0:
        ws12.column_dimensions[get_column_letter(2 + c_idx)].width = 6
r += 1

gantt_data = [
    ("Phase 1", [1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
    ("Phase 2", [0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0]),
    ("Phase 3", [0, 0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0]),
    ("Phase 4", [0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0]),
    ("Phase 5", [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1]),
    ("Testing", [0, 0, 2, 0, 0, 2, 0, 0, 2, 0, 2, 2]),
    ("Training", [0, 0, 0, 0, 0, 0, 0, 0, 0, 3, 3, 3]),
]

phase_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
test_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
train_fill = PatternFill(start_color="006633", end_color="006633", fill_type="solid")

for phase_name, months_active in gantt_data:
    cell = ws12.cell(row=r, column=2, value=phase_name)
    cell.font = bold_font
    cell.border = thin_border
    for m_idx, active in enumerate(months_active):
        cell = ws12.cell(row=r, column=3 + m_idx)
        cell.border = thin_border
        if active == 1:
            cell.fill = phase_fill
            cell.value = "█"
            cell.font = Font(color=WHITE, size=10)
        elif active == 2:
            cell.fill = test_fill
            cell.value = "█"
            cell.font = Font(color=WHITE, size=10)
        elif active == 3:
            cell.fill = train_fill
            cell.value = "█"
            cell.font = Font(color=WHITE, size=10)
        cell.alignment = center_align
    r += 1

# ═══════════════════════════════════════════════════════════════════════
# SHEET 13: PRICING
# ═══════════════════════════════════════════════════════════════════════
ws13 = wb.create_sheet("Pricing")
ws13.sheet_properties.tabColor = DARK_GREEN

for col in range(1, 6):
    ws13.column_dimensions[get_column_letter(col)].width = [5, 18, 15, 35, 15][col - 1]

r = 2
r = write_title(ws13, r, 2, "PRICING & LICENSING", merge_end_col=5)
r += 1

pricing_data = [
    ("Starter", "Up to 50 beds", "Core EMR, OPD, Pharmacy, Lab, Billing", "Custom Quote"),
    ("Professional", "50 – 200 beds", "All Starter + IPD, ER, OT, HR, Analytics", "Custom Quote"),
    ("Enterprise", "200+ beds", "Full suite + PACS, API access, custom dev", "Custom Quote"),
    ("Cloud Hosted", "Any size", "SaaS model — monthly per-user pricing", "Custom Quote"),
    ("On-Premise", "Any size", "Self-hosted with perpetual license", "Custom Quote"),
]
r = write_table(
    ws13, r,
    ["Plan", "Hospital Size", "Includes", "Price"],
    pricing_data,
    col_start=2,
    col_widths=[18, 15, 35, 15],
)
r += 1

r = write_paragraph(
    ws13, r, 2,
    "All plans include: implementation support, data migration, user training (on-site + virtual), "
    "12-month warranty, and technical documentation.",
    merge_end_col=5,
)
r += 1

# Support
r = write_subtitle(ws13, r, 2, "Support & Maintenance", merge_end_col=5)
r += 1

support_data = [
    ("24/7 Help Desk", "Round-the-clock support via phone, email, and chat"),
    ("Dedicated Account Manager", "Single point of contact for your hospital"),
    ("Remote Support", "Screen-sharing troubleshooting and guidance"),
    ("On-Site Support", "Available for critical issues and installations"),
    ("Regular Updates", "Security patches, feature updates, bug fixes"),
    ("Annual Health Check", "System performance review and optimization"),
    ("Training Sessions", "Quarterly refresher training for staff"),
    ("Knowledge Base", "Online documentation and video tutorials"),
    ("SLA Guarantee", "99.9% uptime, <1hr response for critical issues"),
]
r = write_table(ws13, r, ["Service", "Description"], support_data, col_start=2, col_widths=[25, 55])

# ═══════════════════════════════════════════════════════════════════════
# SHEET 14: CHARTS (Dashboard Preview)
# ═══════════════════════════════════════════════════════════════════════
ws14 = wb.create_sheet("Dashboard Preview")
ws14.sheet_properties.tabColor = "003399"

for col in range(1, 8):
    ws14.column_dimensions[get_column_letter(col)].width = 15

r = 2
r = write_title(ws14, r, 1, "DASHBOARD PREVIEW — SAMPLE ANALYTICS", merge_end_col=7)
r += 2

# Sample data for charts
# Department Revenue
ws14.cell(row=r, column=1, value="Department").font = bold_font
ws14.cell(row=r, column=2, value="Monthly Revenue ($)").font = bold_font
dept_data = [
    ("OPD", 45000), ("Surgery", 85000), ("Laboratory", 32000),
    ("Pharmacy", 55000), ("Radiology", 28000), ("Emergency", 38000),
    ("IPD (Wards)", 72000),
]
for i, (dept, rev) in enumerate(dept_data):
    ws14.cell(row=r + 1 + i, column=1, value=dept)
    ws14.cell(row=r + 1 + i, column=2, value=rev)

# Bar chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Monthly Revenue by Department"
chart1.y_axis.title = "Revenue ($)"
chart1.x_axis.title = "Department"
chart1.style = 10
data_ref = Reference(ws14, min_col=2, min_row=r, max_row=r + len(dept_data))
cats_ref = Reference(ws14, min_col=1, min_row=r + 1, max_row=r + len(dept_data))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.width = 20
chart1.height = 12
ws14.add_chart(chart1, f"D{r}")
r += len(dept_data) + 2

# Pie chart — Patient Distribution
pie_r = r + 12
ws14.cell(row=pie_r, column=1, value="Category").font = bold_font
ws14.cell(row=pie_r, column=2, value="Patients").font = bold_font
patient_dist = [
    ("OPD", 450), ("Emergency", 80), ("IPD", 120),
    ("Surgery", 45), ("Maternity", 30),
]
for i, (cat, count) in enumerate(patient_dist):
    ws14.cell(row=pie_r + 1 + i, column=1, value=cat)
    ws14.cell(row=pie_r + 1 + i, column=2, value=count)

pie = PieChart()
pie.title = "Patient Distribution by Department"
pie.style = 10
pie_data = Reference(ws14, min_col=2, min_row=pie_r, max_row=pie_r + len(patient_dist))
pie_cats = Reference(ws14, min_col=1, min_row=pie_r + 1, max_row=pie_r + len(patient_dist))
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.width = 16
pie.height = 12
ws14.add_chart(pie, f"D{pie_r}")

# KPI data
kpi_r = pie_r + len(patient_dist) + 15
ws14.cell(row=kpi_r, column=1, value="KPI").font = bold_font
ws14.cell(row=kpi_r, column=2, value="Current").font = bold_font
ws14.cell(row=kpi_r, column=3, value="Target").font = bold_font
ws14.cell(row=kpi_r, column=4, value="Status").font = bold_font
kpis = [
    ("Bed Occupancy Rate", "78%", "85%", "🟡 Improving"),
    ("Average Wait Time (OPD)", "18 min", "15 min", "🟡 Improving"),
    ("Lab TAT (Routine)", "2.5 hrs", "2 hrs", "🟡 Near Target"),
    ("Patient Satisfaction", "4.2/5", "4.5/5", "🟢 Good"),
    ("Revenue Collection Rate", "92%", "95%", "🟢 Good"),
    ("Insurance Claim Cycle", "18 days", "14 days", "🟡 Improving"),
    ("ER Door-to-Doctor", "8 min", "10 min", "🟢 Exceeding"),
    ("Readmission Rate (30d)", "4.2%", "<5%", "🟢 Good"),
]
for i, (kpi, current, target, status) in enumerate(kpis):
    row_num = kpi_r + 1 + i
    ws14.cell(row=row_num, column=1, value=kpi).font = normal_font
    ws14.cell(row=row_num, column=2, value=current).font = normal_font
    ws14.cell(row=row_num, column=3, value=target).font = normal_font
    ws14.cell(row=row_num, column=4, value=status).font = normal_font
    for c in range(1, 5):
        ws14.cell(row=row_num, column=c).border = thin_border
    ws14.cell(row=row_num, column=c).alignment = wrap_align

# Format KPI headers
for c in range(1, 5):
    ws14.cell(row=kpi_r, column=c).fill = header_fill
    ws14.cell(row=kpi_r, column=c).font = header_font
    ws14.cell(row=kpi_r, column=c).border = thin_border

# ═══════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════
output_path = os.path.join(OUTPUT_DIR, "Hospital_EMR_Pitch.xlsx")
wb.save(output_path)
print(f"✅ Excel workbook saved to: {output_path}")
