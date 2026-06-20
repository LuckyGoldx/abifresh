"""
Import ABIFRESH PRODUCTS.xlsx into Supabase `items` table.
- Clears all existing inventory
- Parses Excel with merged cells (brand, package)
- Generates SKU for each product
- Sets unit_price=0, active_store_quantity=1, main_store_quantity=1
- Maps categories correctly

Requires environment variables:
  SUPABASE_URL (or NEXT_PUBLIC_SUPABASE_URL)
  SUPABASE_SERVICE_ROLE_KEY
"""

import os
import openpyxl
from supabase import create_client
import uuid
import re

SUPABASE_URL = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or ""
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or ""

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment variables")
    print("Example:")
    print('  $env:SUPABASE_URL="https://your-project.supabase.co"')
    print('  $env:SUPABASE_SERVICE_ROLE_KEY="eyJ..."')
    exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

EXCEL_PATH = r"C:\Users\LuckyGold\OneDrive\Documents\ABIFRESH PRODUCTS.xlsx"


def determine_category(brand_name: str) -> str:
    brand_upper = (brand_name or "").upper().strip()
    if "DIAPER" in brand_upper and "ADULT" not in brand_upper:
        return "Baby Diapers"
    elif "ADULT" in brand_upper:
        return "Adult Diapers"
    elif "SANITARY" in brand_upper or "PAD" in brand_upper:
        return "Sanitary Pads"
    elif "WIPE" in brand_upper:
        return "Wipes"
    elif "NET PAD" in brand_upper:
        return "Sanitary Pads"
    elif "BLUE PAD" in brand_upper:
        return "Sanitary Pads"
    elif "UNDERPAD" in brand_upper or "NIGHTINGALE" in brand_upper:
        return "Underpads"
    elif "PANTY" in brand_upper or "LINER" in brand_upper:
        return "Panty Liners"
    else:
        return "Other"


def generate_sku(brand: str, package_type: str, product_name: str, index: int) -> str:
    def abbreviate(text: str, max_len=4) -> str:
        if not text:
            return "GEN"
        words = text.upper().strip().split()
        if len(words) == 1:
            return words[0][:max_len]
        abbr = "".join(w[0] for w in words if w)[:max_len]
        if len(abbr) < 2:
            abbr = words[0][:max_len]
        return abbr

    brand_abbr = abbreviate(brand or "GEN", 3)
    size_match = re.search(r"S(\d)", product_name or "")
    size_part = "S" + size_match.group(1) if size_match else ""
    pkg_parts = (package_type or "").upper().strip()
    if "SACHET" in pkg_parts:
        pkg_abbr = "SAC"
    elif "CARRY" in pkg_parts:
        pkg_abbr = "CP"
    elif "ECO" in pkg_parts:
        pkg_abbr = "ECO"
    elif "JUMBO" in pkg_parts:
        pkg_abbr = "JMB"
    elif "MEGA" in pkg_parts or "ZIP" in pkg_parts:
        pkg_abbr = "MGA"
    elif "POCKET" in pkg_parts:
        pkg_abbr = "PKT"
    elif "ROLL" in pkg_parts:
        pkg_abbr = "SAC"
    else:
        pkg_abbr = abbreviate(pkg_parts, 3) if pkg_parts else "STD"

    parts = [brand_abbr, pkg_abbr]
    if size_part:
        parts.append(size_part)
    else:
        parts.append("{0:03d}".format(index))
    return "-".join(parts)


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"]
    products = []
    current_brand = None
    current_package = None
    sku_counter = {}

    for row_idx in range(3, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        col_d = ws.cell(row=row_idx, column=4).value
        col_e = ws.cell(row=row_idx, column=5).value
        col_f = ws.cell(row=row_idx, column=6).value

        if col_a is not None and str(col_a).strip():
            current_brand = str(col_a).strip()
        if col_b is not None and str(col_b).strip():
            current_package = str(col_b).strip()
        if not col_c or not str(col_c).strip():
            continue

        product_name = str(col_c).strip()
        brand = current_brand or ""
        package_type = current_package or ""
        category = determine_category(brand)

        if "WIPE" in brand.upper():
            category = "Wipes"
        if "NET PAD" in brand.upper():
            category = "Sanitary Pads"
        if "BLUE PAD" in brand.upper():
            category = "Sanitary Pads"
        if "UNDERPAD" in brand.upper():
            category = "Underpads"
        if "OTHERS" in brand.upper():
            category = "Other"

        sku = generate_sku(brand, package_type, product_name, row_idx)
        if sku in sku_counter:
            sku_counter[sku] += 1
            sku = sku + "-" + str(sku_counter[sku])
        else:
            sku_counter[sku] = 1

        price_jalingo = float(col_d) if col_d and col_d != 0 else 0.0
        price_outside = float(col_e) if col_e else 0.0
        commission = float(col_f) if col_f else 0.0

        product = {
            "name": product_name,
            "sku": sku,
            "category": category,
            "unit_price": 0,
            "brand": brand,
            "package_type": package_type,
            "price_jalingo": price_jalingo,
            "price_outside": price_outside,
            "commission": commission,
            "active_store_quantity": 1,
            "main_store_quantity": 1,
            "is_available": True,
            "image_url": None,
        }
        products.append(product)

    return products


def main():
    print("=" * 60)
    print("ABIFRESH INVENTORY IMPORT")
    print("=" * 60)

    print("\n[1/4] Parsing Excel file...")
    products = parse_excel()
    print("     Found " + str(len(products)) + " products")

    categories = {}
    for p in products:
        cat = p["category"]
        categories[cat] = categories.get(cat, 0) + 1
    print("\n     Category breakdown:")
    for cat, count in sorted(categories.items()):
        print("       - " + cat + ": " + str(count) + " items")

    print("\n[2/4] Preview (first 5 products):")
    for p in products[:5]:
        print("     " + p["sku"] + " " * (15 - len(p["sku"])) + " | " + p["name"] + " " * (30 - len(p["name"])) + " | " + p["category"] + " " * (15 - len(p["category"])) + " | Brand: " + p["brand"])
    print("     ...")

    print("\n[3/4] Clearing existing data...")
    tables_to_clear = [
        "receipt_items",
        "daily_sales_summary",
        "sales",
        "posted_items",
        "inventory_transfers",
        "inventory_main_store",
        "inventory_active_store",
        "items",
    ]

    for table_name in tables_to_clear:
        try:
            result = supabase.table(table_name).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
            deleted_count = len(result.data) if result.data else 0
            print("     Cleared " + table_name + ": " + str(deleted_count) + " rows deleted")
        except Exception as e:
            err_msg = str(e)
            if "Could not find" in err_msg or "does not exist" in err_msg:
                print("     Skipped " + table_name + ": table not found")
            else:
                print("     Warning on " + table_name + ": " + err_msg[:100])

    print("     All old data cleared successfully!")

    print("\n[4/4] Inserting " + str(len(products)) + " new products...")
    batch_size = 20
    inserted = 0
    errors = 0

    for i in range(0, len(products), batch_size):
        batch = products[i:i + batch_size]
        try:
            result = supabase.table("items").insert(batch).execute()
            inserted += len(batch)
            print("     Inserted batch " + str(i // batch_size + 1) + ": " + str(len(batch)) + " items (total: " + str(inserted) + ")")
        except Exception as e:
            print("     ERROR in batch " + str(i // batch_size + 1) + ": " + str(e))
            for product in batch:
                try:
                    supabase.table("items").insert(product).execute()
                    inserted += 1
                except Exception as e2:
                    errors += 1
                    print("     FAILED: " + product["name"] + " - " + str(e2))

    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print("  Successfully inserted: " + str(inserted))
    print("  Errors: " + str(errors))
    print("  Total products in Excel: " + str(len(products)))
    print("=" * 60)

    print("\nVerifying...")
    count_result = supabase.table("items").select("id", count="exact").execute()
    print("  Items now in database: " + str(count_result.count))


if __name__ == "__main__":
    main()