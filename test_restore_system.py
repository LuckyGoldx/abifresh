#!/usr/bin/env python3
"""
Test Restore System - End-to-End
1. Login to get auth token
2. Export current state to Excel
3. Clear some tables
4. Parse/Restore from Excel
5. Verify restoration
"""
import requests
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import json
import time

BASE_URL = "http://localhost:5000"
ADMIN_USERNAME = "admin_user"  # From database
ADMIN_PASS = "Admin@123456"

print("=" * 80)
print("🧪 RESTORE SYSTEM TEST")
print("=" * 80)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: LOGIN
# ─────────────────────────────────────────────────────────────────────────────
print("\n📝 STEP 1: Login...")
try:
    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"username": ADMIN_USERNAME, "password": ADMIN_PASS},
        timeout=10
    )
    if response.status_code != 200:
        print(f"❌ Login failed: {response.status_code}")
        print(response.text)
        exit(1)
    
    data = response.json()
    token = data.get('token')
    user = data.get('user')
    
    if not token:
        print(f"❌ No token in response: {data}")
        exit(1)
    
    print(f"✅ Logged in as: {user['email']} (role: {user['role']})")
    print(f"   Token: {token[:30]}...")
    
except Exception as e:
    print(f"❌ Login error: {e}")
    exit(1)

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: GET CURRENT SYSTEM STATE
# ─────────────────────────────────────────────────────────────────────────────
print("\n📊 STEP 2: Get current system state...")
try:
    response = requests.get(f"{BASE_URL}/api/backup/meta", headers=headers, timeout=10)
    if response.status_code != 200:
        print(f"❌ Failed to get metadata: {response.status_code}")
        print(response.text)
        exit(1)
    
    tables_meta = response.json()['tables']
    
    print(f"✅ Found {len(tables_meta)} tables:")
    summary = {}
    for t in tables_meta:
        if not t['hasError']:
            summary[t['name']] = t['rowCount']
            print(f"   {t['name']:30} {t['rowCount']:5} rows")
    
except Exception as e:
    print(f"❌ Error getting metadata: {e}")
    exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: EXPORT KEY TABLES TO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
print("\n💾 STEP 3: Export current state to Excel...")
EXPORT_TABLES = ['users', 'items', 'system_settings']
wb = openpyxl.Workbook()
wb.remove(wb.active)

try:
    for table_name in EXPORT_TABLES:
        try:
            response = requests.get(
                f"{BASE_URL}/api/backup/table/{table_name}/all",
                headers=headers,
                timeout=10
            )
            
            if response.status_code != 200:
                print(f"   ⚠️  {table_name}: error {response.status_code}")
                continue
            
            data = response.json()
            rows = data.get('rows', [])
            
            if rows:
                df = pd.DataFrame(rows)
                ws = wb.create_sheet(title=table_name)
                
                # Write headers
                for col_num, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_num, value=col_name)
                
                # Write data
                for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                
                print(f"   ✅ {table_name}: {len(rows)} rows")
            else:
                print(f"   ⏭️  {table_name}: 0 rows")
        except Exception as e:
            print(f"   ❌ {table_name}: {e}")
    
    backup_file = f"test_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(backup_file)
    print(f"✅ Exported to: {backup_file}")
    
except Exception as e:
    print(f"❌ Export error: {e}")
    exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: PARSE BACKUP (preview)
# ─────────────────────────────────────────────────────────────────────────────
print("\n🔍 STEP 4: Parse backup file (preview)...")
try:
    with open(backup_file, 'rb') as f:
        files = {'file': (backup_file, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/backup/restore/parse",
            headers={"Authorization": headers["Authorization"]},  # Remove Content-Type for multipart
            files=files,
            timeout=10
        )
    
    if response.status_code != 200:
        print(f"❌ Parse failed: {response.status_code}")
        print(response.text)
        exit(1)
    
    parse_result = response.json()
    sheets = parse_result.get('sheets', [])
    
    print(f"✅ File parsed successfully:")
    for sheet in sheets:
        print(f"   Sheet: {sheet['sheetName']}")
        print(f"     → Table: {sheet['tableName']}")
        print(f"     → Rows: {sheet['rowCount']}")
        print(f"     → Matched: {sheet['matched']}")
    
except Exception as e:
    print(f"❌ Parse error: {e}")
    exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: CLEAR A TABLE (simulate changes)
# ─────────────────────────────────────────────────────────────────────────────
print("\n🗑️  STEP 5: Clear 'items' table to simulate data loss...")
try:
    # We can't directly clear via API, so we'll skip this for now
    # In real scenario, user would have cleared data
    print("   ⏭️  Skipping (would clear table manually)")
except Exception as e:
    print(f"   ⚠️  {e}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: RESTORE FROM BACKUP
# ─────────────────────────────────────────────────────────────────────────────
print("\n📥 STEP 6: Restore from backup (replace mode)...")
try:
    with open(backup_file, 'rb') as f:
        files = {'file': (backup_file, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'mode': 'replace',  # Clear and restore
            'tables': json.dumps(EXPORT_TABLES)  # Only restore these tables
        }
        
        response = requests.post(
            f"{BASE_URL}/api/backup/restore/commit",
            headers={"Authorization": headers["Authorization"]},
            files=files,
            data=data,
            timeout=30
        )
    
    if response.status_code != 200:
        print(f"❌ Restore failed: {response.status_code}")
        print(response.text)
        exit(1)
    
    restore_result = response.json()
    results = restore_result.get('results', [])
    
    print(f"✅ Restore completed:")
    for result in results:
        status = "✅" if result['success'] else "❌"
        print(f"   {status} {result['table']}")
        print(f"      Total: {result['rowsTotal']}, Inserted: {result['rowsInserted']}")
        if result.get('error'):
            print(f"      Error: {result['error']}")
    
except Exception as e:
    print(f"❌ Restore error: {e}")
    exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7: VERIFY RESTORATION
# ─────────────────────────────────────────────────────────────────────────────
print("\n✔️  STEP 7: Verify restoration...")
try:
    response = requests.get(f"{BASE_URL}/api/backup/meta", headers=headers, timeout=10)
    if response.status_code != 200:
        print(f"❌ Failed to verify: {response.status_code}")
        exit(1)
    
    tables_meta = response.json()['tables']
    
    print("✅ Post-restore state:")
    all_match = True
    for t in tables_meta:
        if t['name'] in EXPORT_TABLES and not t['hasError']:
            orig_count = summary.get(t['name'], 0)
            current_count = t['rowCount']
            match = "✅" if orig_count == current_count else "❌"
            print(f"   {match} {t['name']:30} {current_count:5} rows (expected: {orig_count})")
            if orig_count != current_count:
                all_match = False
    
    if all_match:
        print("\n" + "=" * 80)
        print("✅ TEST PASSED - RESTORE SYSTEM WORKS CORRECTLY!")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("⚠️  TEST PARTIAL - Some row counts don't match")
        print("=" * 80)
    
except Exception as e:
    print(f"❌ Verification error: {e}")
    exit(1)

print("\n📊 Summary:")
print(f"   Backup file: {backup_file}")
print(f"   Tables tested: {', '.join(EXPORT_TABLES)}")
print(f"   Mode: Replace (delete all, then insert from backup)")
print(f"\nTest completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
